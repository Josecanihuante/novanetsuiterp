"""
Servicio NetSuite — Importación y parsing de Excel exportado desde NetSuite.

Columnas esperadas (en cualquier orden):
    Account, Account Name, Type, Date, Document Number,
    Name, Memo, Debit, Credit, Amount, Currency,
    Subsidiary, Department, Class
"""
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

COLUMNAS: list[str] = [
    "Account",
    "Account Name",
    "Type",
    "Date",
    "Document Number",
    "Name",
    "Memo",
    "Debit",
    "Credit",
    "Amount",
    "Currency",
    "Subsidiary",
    "Department",
    "Class",
]

TIPO_MAP: dict[str, str] = {
    "Income": "income",
    "Expense": "expense",
    "Asset": "asset",
    "Liability": "liability",
    "Equity": "equity",
}


def parse_netsuite_excel(file_bytes: bytes) -> dict[str, Any]:
    """
    Parsea un archivo Excel exportado desde NetSuite.

    Parámetros
    ----------
    file_bytes : contenido binario del archivo .xlsx

    Retorna
    -------
    {
        "valid_rows": [...],    # filas correctamente parseadas
        "errors":     [...],    # errores por fila/columna
        "total_rows": int,
        "valid_count": int,
        "error_count": int,
    }

    Lanza
    -----
    ValueError si faltan columnas requeridas en el encabezado.
    """
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    # ── Validar encabezados ───────────────────────────────────────────────────
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    faltantes = [col for col in COLUMNAS if col not in headers]
    if faltantes:
        raise ValueError(f"Columnas faltantes en el Excel: {faltantes}")

    valid_rows: list[dict] = []
    errors: list[dict] = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        r = dict(zip(headers, row))

        # Saltar filas completamente vacías
        if all(v is None for v in r.values()):
            continue

        row_errors: list[dict] = []

        # ── Validar Date (MM/DD/YYYY o datetime de Excel) ─────────────────────
        fecha = None
        try:
            raw_date = r["Date"]
            if isinstance(raw_date, datetime):
                fecha = raw_date.date()
            else:
                fecha = datetime.strptime(str(raw_date), "%m/%d/%Y").date()
        except Exception:
            row_errors.append({
                "row": i,
                "column": "Date",
                "message": "Formato de fecha inválido. Use MM/DD/YYYY",
                "value": r.get("Date"),
            })

        # ── Validar Debit ──────────────────────────────────────────────────────
        debito = None
        try:
            debito = float(r["Debit"] or 0)
        except Exception:
            row_errors.append({
                "row": i,
                "column": "Debit",
                "message": "El valor de Debit debe ser numérico",
                "value": r.get("Debit"),
            })

        # ── Validar Credit ─────────────────────────────────────────────────────
        credito = None
        try:
            credito = float(r["Credit"] or 0)
        except Exception:
            row_errors.append({
                "row": i,
                "column": "Credit",
                "message": "El valor de Credit debe ser numérico",
                "value": r.get("Credit"),
            })

        if row_errors:
            errors.extend(row_errors)
        else:
            valid_rows.append({
                "account_code": str(r["Account"]) if r["Account"] is not None else None,
                "account_name": r["Account Name"],
                "account_type": TIPO_MAP.get(str(r["Type"]) if r["Type"] else "", "expense"),
                "date": fecha,
                "document_number": str(r["Document Number"]) if r["Document Number"] is not None else None,
                "entity_name": r["Name"],
                "description": r["Memo"],
                "debit": debito,
                "credit": credito,
                "amount": r.get("Amount"),
                "currency": r.get("Currency", "CLP"),
                "subsidiary": r.get("Subsidiary"),
                "department": r.get("Department"),
                "class_field": r.get("Class"),
            })

    wb.close()

    return {
        "valid_rows": valid_rows,
        "errors": errors,
        "total_rows": len(valid_rows) + len(errors),
        "valid_count": len(valid_rows),
        "error_count": len(errors),
    }
